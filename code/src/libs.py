from cv2 import log
from statsmodels.stats.anova import anova_lm
from statsmodels.formula.api import ols
from scipy import stats
from base64 import encode
from encodings import utf_8
import encodings
from operator import index
import os
from pathlib import Path
import re
from statistics import mode
from this import d
from time import time
import numpy as np
import pandas as pd
import openpyxl
import yaml

RESULTSPATH = './results.yml'

def Log(*content:str):
    '''整齐打印字符串or分割行
    Args：
        *contend(str) : 不输入参数时，则打印分割行；输入字符串参数时则打印该字符串
    '''
    if len(content) == 0:
        print('=·=·=·=·=·=·=·=·=·=·=·=·=·=·=·=·=·=·=·=·=·=·=·=·=')
    else:
        content = content[0]
        default_filler = '~·'
        left_filler = ''
        for i in range(int((50-len(content))/4)):
            left_filler += default_filler
        print(left_filler + content + '·' + left_filler + '~')

def GetFilesPaths(parent_path:str, file_type:str, log = True):
    '''根据指定文件类型，在父级目录下查找指定类型的文件
    内部使用os.listdir不支持嵌套文件夹
    Args:
        parent_path (str): 父级目录地址，相对And绝对，以/结尾
        file_type (str): 文件类型

    Returns:
        list: 查找结果，文件路径list
        0: 输入路径不是一个文件夹
        -1： 输入路径内无文件
    '''
    if os.path.isdir(parent_path):
        res = []
        file_list = os.listdir(parent_path)
        if file_list:
            for file in file_list:
                if file.endswith(file_type):
                    file_path = os.path.join(parent_path, file)  # 拼接成路径
                    res.append(file_path)
            if log:
                Log('File Searching Done')
                return res
            else:
                return res
        else:
            Log('No File in This Folder')
            return -1
    else:
        Log("Not a Folder Path")
        return 0
    
def LoadTxtFile(file_path:str, encoding = utf_8):
    '''By given txt file path, loading txt file content into list of string

    Args:
        file_path (str): txt file path

    Returns:
        lsit: content of txt file in str
    '''
    if os.path.exists(file_path):
        file_contents = []
        for line in open(file_path, 'r', encoding=encoding):
           file_contents.append(CutString(line))
        return file_contents
    else:
        print('File Doesn\'t Exist!')
        return None

def CutString(str_in:str):
    '''cut string and return list of float

    Args:
        str_in (str): str in

    Returns:
        list: float value in the str 
    '''
    return list(map(float, str_in.split('    ', 2)))

def ConvertRows2Columns(data:list):
    '''pick up the columns item and make a new list of these columns items

    Args:
        data (list): origin data that contains [time, x, y] , a 2-d list goes like:
        [   [time, x, y]
            [time, x, y]
            [time, x, y]
            ...
            [time, x, y]
            [time, x, y]
        ]

    Returns:
        list: pick up all the times xs and ys make them 3 individual list, goes like:
        [
            [time1, time2, time3, ... time5]
            [x1, x2,x3, ... x5]
            [y1, y2,y3, ... y5]
        ]
    '''
    time_data = [i[0] for i in data]
    x_data = [i[1] for i in data]
    y_data = [i[2] for i in data]
    return [time_data, x_data, y_data]

def WriteTimeXYData(data1:list):
    '''by given times_x_y list data, write it in a workbook

    Args:
        data1 (list): [[times], [xs], [ys]]

    Returns:
        workbook(openpyxl.Workbook): workbook contains the data
    '''
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Data'
    row_index = 0
    column_index = -2   # because the first if the loop s gonna add the column_index by 3
    time_data = data1[0]
    for index in range(len(time_data)):
        if time_data[index] == 0.1: # when find a time equals to 0.1
            row_index = 1 # sheet row idnex should be reset as 1
            column_index += 3 #  sheet column index should be added by 3
        worksheet.cell(row_index, column_index, data1[0][index])    # time data
        worksheet.cell(row_index, column_index + 1, data1[1][index])    # x
        worksheet.cell(row_index, column_index + 2, data1[2][index])    # y
        row_index += 1
    Log(str(column_index))
    return workbook

def GetEndTimeByTimeColumn(time_series:pd.Series):
    '''在提供的时间序列中查找最后一个非nan时间

    Args:
        time_series (pd.Series): 给定的时间序列

    Returns:
        float: 给定时间序列的最后一个非nan数字
    '''
    tmp = -1
    for i in time_series:
        if i != i:
            return tmp
        else:
            tmp = i
    return time_series[time_series.size - 1]

def GetStartTime(x_series:list):
    '''get the movement start time for single ped by given this ped's movement x pos

    Args:
        x_series (list): ped's movement x positions

    Returns:
        start time(float): given ped's movement start time as a float
    '''
    flag = x_series[0]
    for i in range(len(x_series)):
        if x_series[i] != flag:
            return (i+1) * 0.1

def FixData(origin_data:pd.DataFrame, correct_columns: int, *axis):
    '''fix the given data in case this data has empty columns or descirpion information at the header

    Args:
        origin_data (pd.DataFrame): the origin data readed from the xlsx
        correct_columns (int): the correctly number of columns

    Returns:
        fix_info (dict): a dict contains the fix information, and this would be used as origin the yml input
    '''
    fix_info = {}
    rows, columns = origin_data.shape
    print(origin_data.index)
    reindex = 1
    # row correct
    if origin_data.iloc[0][0] != 0.1: # 如果第一行一个元素不是数字0.1，则进行校正
            origin_data = origin_data.drop([0]) # 删除第一行
            fix_info['Row Correct'] = 'True' # 将是否校正第一行添加到valu
            reindex = 0
    if columns == correct_columns:
        pass
    else:
        origin_data = origin_data.dropna(axis=1, how='all')
        fix_info['Column Correct'] = 'True' # 将是否校正第一行添加到valu
    print(origin_data.index)
    return origin_data, fix_info, reindex

def DataAna(file:str):
    '''data analysis

    Args:
        file (str): the xlsx filepath
        CORRECT_COLUMNS (int, optional): the correct columns number. Defaults to 198.

    Returns:
        CaseName: the file name
        Info: the fix information
    '''
    CaseName = file.split('.xlsx')[0].split('/')[-1]    # get the file name
    Log(CaseName)
    OriginData = pd.read_excel(file, header=None)   # read origin data
    row, column = OriginData.shape
    print(column, int(column/3) * 3)
    OriginData, Info, ReIndex = FixData(OriginData, (column/3)* 3)    # fix data
    row, column = OriginData.shape
    PersonCount = float(column/3)
    Log("Total " + str(column/3) + ' Persons')
    StartTime = []
    EndTime = []
    EndPostion = []
    for i in range(column):
        if i%3 == 0:
            # end time
            single_endtime = GetEndTimeByTimeColumn(OriginData[i])
            EndTime.append(single_endtime)
            StartTime.append(GetStartTime(list(OriginData[i+1])))
            # end position
            EndPostion.append([OriginData[i+1][single_endtime*10-ReIndex], OriginData[i+2][single_endtime*10-ReIndex]])
    # data category
    StartTimeArray = np.array(StartTime)
    EndTimeArray = np.array(EndTime)
    LeftEndTime = EndTimeArray[:33]
    RightEndTime = EndTimeArray[33:]
    LeftStartTime = StartTimeArray[:33]
    RightStartTime = StartTimeArray[33:]
    # info 
    Info['Person'] = PersonCount    # add the person info into the info
    Info['MaxTime'] = {'Total':str(EndTime), "Left":str(EndTimeArray[:33]), "Right":str(EndTimeArray[33:])} # 将每个人的运动时间添加到字典的value
    Info['EndPosition'] = str(EndPostion)
    Info['MeanMaxTime'] = {'Total':float(np.mean(EndTimeArray)), 'Left':float(np.mean(LeftEndTime)), 'Right':float(np.mean(RightEndTime))}
    Info['StdMaxTime'] = {'Total':float(np.std(EndTimeArray)), 'Left':float(np.std(LeftEndTime)), 'Right':float(np.std(RightEndTime))}
    Info['StartTime'] = {'Total':str(StartTime), "Left":str(StartTimeArray[:33]), "Right":str(StartTimeArray[33:])} # 将每个人的运动时间添加到字典的value
    Info['MeanStartTime'] = {'Total':float(np.mean(StartTimeArray)), 'Left':float(np.mean(LeftStartTime)), 'Right':float(np.mean(RightStartTime))}
    Info['StdStartTime'] = {'Total':float(np.std(StartTimeArray)), 'Left':float(np.std(LeftStartTime)), 'Right':float(np.std(RightStartTime))}
    Info['VarStartTime'] = {'Total':float(np.var(StartTimeArray)), 'Left':float(np.var(LeftStartTime)), 'Right':float(np.var(RightStartTime))}
    Log(CaseName + ' Done')
    return CaseName, Info

def LoadYml(filepath:str):
    '''load the yml file by a given path

    Args:
        filepath (str): ynl file path

    Returns:
        file content (dict): the yml file content
    '''
    with open(filepath, 'r', encoding='utf-8') as f:
        return yaml.load(f.read(),Loader=yaml.Loader)

def LogResults(results:dict, CaseNames:list, *Items):
    '''get specific results data from the Data readed from the yml by givben project indexs

    Args:
        results (dict): the data readed from the yml
        CaseNames (list): project index

    Returns:
        results (dict): {case index: data}
    '''
    Log(str(CaseNames))
    LogContent = ''
    LogRes = {}
    # 递归
    for i in CaseNames:
        LogContent= results[str(i)]
        for j in Items:
            LogContent = LogContent[j]
        # print(LogContent) 
        LogRes[i] = LogContent
    Log()
    return LogRes

def CutPostionStr2PositionList(s:str):
    s = s.replace('[', '')
    s = s.replace(' ', '')
    s = s.replace(']', '')
    l1 = s.split(',', 200)
    res = []
    x_pos = l1[::2]
    y_pos = l1[1::2]
    for i in range(len(x_pos)):
        res.append([float(x_pos[i]), float(y_pos[i])])
    return res

def CutTimeStr2TimeList(s:str):
    s = s.strip('[')
    s = s.strip(']')
    l1 = s.split(',', 200)
    return list(map(float, l1))

def JudgeLR(postions:list, CaseMode:int):
    '''passing in a list contains end positions and judge the right or left of every single positon

    Args:
        postions (list): the origin list of end positions
        CaseMode (int): the current positions list mode:1-left1_right2, 2-left2_right1, 0-both unknown or both known

    Returns:
        _type_: _description_
    '''
    count_l = 0
    for i in postions:
        if i[0] < 0:
            count_l += 1
    if CaseMode == 1:
        return {'Unknown':count_l, 'Known': len(postions) - count_l}
    elif CaseMode == 2:
        return {'Known':count_l, 'Unknown': len(postions) - count_l}
    else:
        return {'Left':count_l, 'Right': len(postions) - count_l}

def DumpYml(path:str, content:dict):
    with open(path, "w", encoding="utf-8") as f:
        yaml.dump(content, f)

def IQRDetectOutliers(timelist:list, range:list):
    '''根据指定的Interquartile Rnage筛选出给定数据中的Outliers

    Args:
        df (pd.DataFrame): 原始数据
        IQR (list): IQR的范围
    Return:
        df_out (pd.DataFrame): 去除异常值后的数据
    '''
    df = pd.DataFrame(timelist)
    Q1 = df.quantile(range[0])
    Q3 = df.quantile(range[1])
    IQR = Q3- Q1
    
    return df[~((df < (Q1 - 1.5*IQR))|(df>(Q3 + 1.5*IQR))).any(axis=1)]

def expend(length:int, item):
    res = []
    for i in range(length):
        res.append(item)
    return res
def generatedfanalysisdata(data:dict, columns:list,hue:dict):
    '''传入字典数据,据字典数据产生用于dataframe分析的数据

    Args:
        data (dict): {key(index):value(list(int))}

    Returns:
        list: [[index], [value]]
    '''
    #   生成反向的hue字典
    reverse_hue = {}
    for key, value in hue.items():
        for i in list(value):
            reverse_hue[i] = key
    values = []
    indexes = []
    hues = []
    for key, value in data.items():
        values += value
        tmp_len = len(value)
        indexes += expend(tmp_len, key)
        hues += expend(tmp_len, reverse_hue[key])
    if len(values) == len(indexes):
        return pd.DataFrame({columns[0]:values, columns[1]:indexes, 'hue':hues})
    else:
        Log('Something Wrong')
        return None
def generatedfanalysisdata_sim(data:dict, columns:list):
    '''传入字典数据,据字典数据产生用于dataframe分析的数据

    Args:
        data (dict): {key(index):value(list(int))}

    Returns:
        list: [[index], [value]]
    '''
   
    values = []
    indexes = []
    for key, value in data.items():
        values += value
        tmp_len = len(value)
        indexes += expend(tmp_len, key)
       
    if len(values) == len(indexes):
        return pd.DataFrame({columns[0]:values, columns[1]:indexes})
    else:
        Log('Something Wrong')
        return None



def anova_oneway(data:dict):
    '''单因素检验

    Args:
        data (dict): _description_
    '''
    #   检查方差是否相等
    datavalue = list(data.values())
    (W,p) = stats.levene(datavalue[0], datavalue[1], datavalue[2])
    if p >= 0.05:
        Log('variances are equal')
    else:
        Log('variances are not equal')
    #   进行实际的oneway检验
    F_statistic, pVal = stats.f_oneway(datavalue[0], datavalue[1], datavalue[2])
    print((F_statistic, pVal))
    if pVal < 0.05:
        print('One of the groups is significantly different.')
    data = generatedfanalysisdata_sim(data, ['value', 'index'])
    model = ols('value ~ C(index)', data).fit()
    anovaResults = anova_lm(model)
    print(anovaResults)


def ttest_2(l1:list, l2:list):
    Log('t test begins')
    args = [l1,l2]
    w, p = stats.levene(*args)
    print(np.array(l1).mean(), np.array(l1).std())
    print(np.array(l2).mean(), np.array(l2).std())
    if p >= 0.05:
        Log('方差相等')
        print(stats.ttest_ind(l1,l2,equal_var=True))
    else:
        Log('方差不等')
        print(stats.ttest_ind(l1,l2,equal_var=False))
    Log('t test ends')
def loopttest(l1:list, comparelist:list):
    for i in comparelist:
        ttest_2(l1, i)

if __name__ == "__main__":
    # 
    Log('Main function Running')


